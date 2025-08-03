import sys
import openpyxl

def validate_xlsx(file_path):
    """
    Opens an XLSX file and prints its contents and metadata for debugging.
    - Lists all sheet names and their indices.
    - Dumps the raw value of every cell in each sheet.
    - Reports any merged cells found in each sheet.
    """
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
    except FileNotFoundError:
        print(f"Error: The file '{file_path}' was not found.")
        return
    except Exception as e:
        print(f"An error occurred while trying to open the file: {e}")
        return

    print(f"--- Validating: {file_path} ---\n")

    # 1. List all sheets
    print("--- Available Sheets ---")
    if not workbook.sheetnames:
        print("No sheets found in this workbook.")
        return
        
    for i, sheet_name in enumerate(workbook.sheetnames):
        print(f"Index {i + 1}: '{sheet_name}'")
    print("-" * 26 + "\n")

    # 2. Dump raw data and check for merged cells for each sheet
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"--- Content of sheet: '{sheet_name}' ---")

        # Check for merged cells
        if sheet.merged_cells:
            print("\n[WARNING] Merged cells found! This can cause parsing issues.")
            for merged_range in sheet.merged_cells:
                print(f"- Merged range: {merged_range}")
            print()

        # Dump raw cell data
        if sheet.max_row == 0:
            print("(This sheet is empty)")
        else:
            for row_idx, row in enumerate(sheet.iter_rows()):
                row_values = [cell.value if cell.value is not None else "" for cell in row]
                print(f"Row {row_idx + 1}: {row_values}")
        
        print("-" * (26 + len(sheet_name)) + "\n")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python validate_xlsx.py <path_to_your_excel_file.xlsx>")
        sys.exit(1)
    
    file_to_validate = sys.argv[1]
    validate_xlsx(file_to_validate)