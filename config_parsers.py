import openpyxl
import csv
from itertools import zip_longest

def parse_nodes_from_csv(file_path):
    nodes = []
    config_headers = {
        'nodename', 'protocol', 'ip_address', 'login_id', 
        'login_password', 'additional_command_1', 'additional_command_2'
    }

    try:
        with open(file_path, 'r', newline='', encoding='utf-8-sig') as csvfile:
            reader = list(csv.reader(csvfile))
            if not reader:
                return []
            
            transposed_data = list(zip_longest(*reader, fillvalue=''))
            headers = [h.strip() for h in transposed_data[0]]

            for i in range(1, len(transposed_data)):
                node_info = {"commands": []}
                node_column = transposed_data[i]
                commands_ended = False

                for j, header in enumerate(headers):
                    value = node_column[j].strip() if j < len(node_column) else ""

                    if header not in config_headers:
                        if not value:
                            commands_ended = True
                        
                        if not commands_ended:
                            node_info["commands"].append(value)
                    else:
                        node_info[header] = value
                
                if node_info.get('nodename'):
                    nodes.append(node_info)

    except FileNotFoundError:
        print(f"Error: The file {file_path} was not found.")
        return None
    except Exception as e:
        print(f"An error occurred while parsing the CSV: {e}")
        return None
    return nodes

def parse_nodes_from_excel(file_path, sheet_name=1):
    nodes = []
    config_headers = {
        'nodename', 'protocol', 'ip_address', 'login_id',
        'login_password', 'additional_command_1', 'additional_command_2'
    }

    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        if isinstance(sheet_name, int):
            if 1 <= sheet_name <= len(workbook.sheetnames):
                sheet = workbook.worksheets[sheet_name - 1]
            else:
                raise ValueError(f"Sheet index {sheet_name} is out of range.")
        else:
            sheet = workbook[sheet_name]

        # Read data into a 2D list (list of rows)
        data_by_rows = []
        for row in sheet.iter_rows():
            data_by_rows.append([cell.value if cell.value is not None else "" for cell in row])

        # Corrected "Fill-Right" Logic
        for row in data_by_rows:
            # The first column (index 0) contains headers and should not be filled.
            # We start the logic from the first data column (index 1).
            if len(row) > 1:
                last_value = row[1]  # Initialize with the first data point
                for i in range(2, len(row)):
                    if str(row[i]).strip() == "":
                        row[i] = last_value
                    else:
                        last_value = row[i]

        # Transpose the processed data to get columns for parsing
        if not data_by_rows or not data_by_rows[0]:
            return []
        transposed_data = list(zip_longest(*data_by_rows, fillvalue=''))

        headers = [str(h).strip() for h in transposed_data[0]]

        for i in range(1, len(transposed_data)):
            node_info = {"commands": []}
            node_column = transposed_data[i]
            commands_ended = False

            for j, header in enumerate(headers):
                value = str(node_column[j]).strip() if j < len(node_column) else ""

                if header not in config_headers:
                    if not value:
                        commands_ended = True
                    
                    if not commands_ended:
                        node_info["commands"].append(value)
                else:
                    node_info[header] = value
            
            if node_info.get('nodename'):
                nodes.append(node_info)

    except FileNotFoundError:
        print(f"Error: The file {file_path} was not found.")
        return None
    except Exception as e:
        print(f"An error occurred while parsing the Excel file: {e}")
        return None
    return nodes